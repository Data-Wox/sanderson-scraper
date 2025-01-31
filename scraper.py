from webdriver import Webdriver
from openpyxl import Workbook, load_workbook
from database import insert_relatorio_0063, verify_relatorio_0063, insert_relatorio_0123, verify_relatorio_0123, verify_relatorio_0053, insert_relatorio_0053, verify_relatorio_0004, insert_relatorio_0004, insert_relatorio_0186, verify_relatorio_0186, add_log

import time
import os

EMPRESA_LINK = ''
EMPRESA_EMAIL = ''
EMPRESA_PASSWORD = ''
EMPRESA_ID = 0

INIT_DATE_0186 = ''
FINAL_DATE_0186 = ''

INIT_DATE_0123 = ''
FINAL_DATE_0123 = ''

INIT_DATE_0053 = ''
FINAL_DATE_0053 = ''

def login(driver):
    global EMPRESA_LINK
    global EMPRESA_EMAIL
    global EMPRESA_PASSWORD

    try:
        driver.get(EMPRESA_LINK)

        time.sleep(5)

        emailForm = driver.execute_script("return document.getElementById('formEmail')")
        passwordForm = driver.execute_script("return document.getElementById('formSenha')")

        emailForm.send_keys(EMPRESA_EMAIL)
        passwordForm.send_keys(EMPRESA_PASSWORD)

        time.sleep(1)

        driver.execute_script("document.getElementsByClassName('matomo-realizar-login')[0].click()")

        time.sleep(5)

        driver.execute_script("document.getElementsByClassName('btn cursor-pointer')[0].click()")

        time.sleep(1)
    except Exception as e:
        add_log('Error when logging in: ' + str(e))

def parse_relatorio_0186(fileName):
    letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

    wb = load_workbook(fileName)
    ws = wb.active

    totalRows = 2

    while ws['A' + str(totalRows)].value != None:
        totalRows += 1

    total_data = []

    def get_right_keys(key):
        if key == 'Data':
            return 'data'
        elif key == 'Comanda':
            return 'comanda'
        elif key == 'Item':
            return 'item'
        elif key == 'Tipo':
            return 'tipo'
        elif key == 'Categoria':
            return 'categoria'
        elif key == 'Profissional':
            return 'profissional'
        elif key == 'Assistente 1':
            return 'assistente_1'
        elif key == 'Assistente 2':
            return 'assistente_2'
        elif key == 'Comissão (%)':
            return 'comissao_percentual'
        elif key == 'Cliente':
            return 'cliente'
        elif key == 'Email':
            return 'email'
        elif key == 'Telefone':
            return 'telefone'
        elif key == 'Celular':
            return 'celular'
        elif key == 'Valor':
            return 'valor'
        elif key == 'Desconto':
            return 'desconto'
        elif key == 'Qtd.':
            return 'quantidade'
        elif key == 'Custo':
            return 'custo'
        elif key == 'Comissão':
            return 'comissao'
        elif key == 'Líquido':
            return 'liquido'
        elif key == 'UA':
            return 'ua'

    for n in range(2, totalRows):
        data = {}
        for i in range(0, 20):
            key = get_right_keys(ws['{}1'.format(letters[i])].value.strip())
            data[key] = ws['{}{}'.format(letters[i], n)].value
        total_data.append(data)

    return total_data

def parse_relatorio_0063(fileName):
    letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

    wb = load_workbook(fileName)
    ws = wb.active

    totalRows = 2

    while ws['A' + str(totalRows)].value != None:
        totalRows += 1

    total_data = []

    def get_right_keys(key):
        if key == 'Cliente':
            return 'cliente'
        if key == 'Pacote':
            return 'pacote'
        if key == 'Serviço':
            return 'servico'
        if key == 'Total':
            return 'total'
        if key == 'Utilizados':
            return 'utilizados'
        if key == 'Validade':
            return 'validade'
        if key == 'Compra':
            return 'compra'

    for n in range(2, totalRows):
        data = {}
        for i in range(0, 7):
            key = get_right_keys(ws['{}1'.format(letters[i])].value.strip())
            data[key] = ws['{}{}'.format(letters[i], n)].value
        total_data.append(data)

    return total_data

def parse_relatorio_0123(fileName):
    letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

    wb = load_workbook(fileName)
    ws = wb.active

    totalRows = 2

    while ws['A' + str(totalRows)].value != None:
        totalRows += 1

    total_data = []

    def get_right_keys(key):
        if key == 'Profissional':
            return 'profissional'
        if key == 'Tipo Contratação':
            return 'tipo_contratacao'
        if key == 'Cargo':
            return 'cargo'
        if key == 'Banco':
            return 'banco'
        if key == 'Agência':
            return 'agencia'
        if key == 'Conta':
            return 'conta'
        if key == 'Faturado':
            return 'faturado'
        if key == 'Rateio Serviços':
            return 'rateio_servicos'
        if key == 'Rateio Produtos':
            return 'rateio_produtos'
        if key == 'Rateio Outros':
            return 'rateio_outros'
        if key == 'Caixinha':
            return 'caixinha'
        if key == 'Descontos':
            return 'descontos'
        if key == 'A pagar':
            return 'a_pagar'
        if key == 'Valor Casa':
            return 'valor_casa'

    for n in range(2, totalRows):
        data = {}
        for i in range(0, 14):
            key = get_right_keys(ws['{}1'.format(letters[i])].value.strip())
            value = str(ws['{}{}'.format(letters[i], n)].value)
            if (key == 'faturado' or key == 'rateio_servicos' or key == 'rateio_produtos' or key == 'rateio_outros' or key == 'caixinha' or key == 'descontos' or key == 'a_pagar' or key == 'valor_casa'):
                data[key] = float(value.replace(',', '.'))
            else: 
                data[key] = value
        total_data.append(data)

    return total_data

def parse_relatorio_0004(fileName):
    letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

    wb = load_workbook(fileName)
    ws = wb.active

    totalRows = 2

    while ws['A' + str(totalRows)].value != None:
        totalRows += 1

    total_data = []

    def get_right_keys(key):
        if key == 'Cliente':
            return 'cliente'
        if key == 'Código':
            return 'codigo'
        if key == 'Aniversário':
            return 'aniversario'
        if key == 'Telefone':
            return 'telefone'
        if key == 'Celular':
            return 'celular'
        if key == 'E-mail':
            return 'email'
        if key == 'Sexo':
            return 'sexo'
        if key == 'Como Conheceu':
            return 'como_conheceu'
        if key == 'CPF':
            return 'cpf'
        if key == 'CEP':
            return 'cep'
        if key == 'Endereço':
            return 'endereco'
        if key == 'Número':
            return 'numero'
        if key == 'Estado':
            return 'estado'
        if key == 'Cidade':
            return 'cidade'
        if key == 'Complemento':
            return 'complemento'
        if key == 'Bairro':
            return 'bairro'
        if key == 'Profissão':
            return 'profissao'
        if key == 'Cadastrado':
            return 'cadastrado'
        if key == 'Obs':
            return 'obs'
        if key == 'RG':
            return 'rg'

    for n in range(2, totalRows):
        data = {}
        for i in range(0, 20):
            key = get_right_keys(ws['{}1'.format(letters[i])].value.strip())
            data[key] = ws['{}{}'.format(letters[i], n)].value
        total_data.append(data)

    return total_data

def parse_relatorio_0053(fileName):
    letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']

    wb = load_workbook(fileName)
    ws = wb.active

    totalRows = 2

    while ws['A' + str(totalRows)].value != None:
        totalRows += 1

    total_data = []

    def get_right_keys(key):
        if key == 'Data Reserva':
            return 'data_reserva'
        if key == 'Hora':
            return 'hora'
        if key == 'Cliente':
            return 'cliente'
        if key == 'Serviço':
            return 'servico'
        if key == 'Valor':
            return 'valor'

    for n in range(2, totalRows):
        data = {}
        for i in range(0, 5):
            key = get_right_keys(ws['{}1'.format(letters[i])].value.strip())
            data[key] = ws['{}{}'.format(letters[i], n)].value
        total_data.append(data)

    return total_data

def relatorio_0186(driver, idEmpresa):
    driver.get('https://admin.avec.beauty/admin/relatorios/0186')
    time.sleep(3)

    driver.execute_script("document.getElementsByClassName('inputVarRel')[0].value = '{}'".format(INIT_DATE_0186))
    driver.execute_script("document.getElementsByClassName('inputVarRel')[1].value = '{}'".format(FINAL_DATE_0186))

    time.sleep(15)

    driver.execute_script("document.getElementsByClassName('btn btn-info hidden-print')[0].click()")

    time.sleep(7)

    driver.execute_script("document.getElementsByClassName('buttons-excel')[0].click()")

    time.sleep(2)

    driver.get('chrome://downloads')
    time.sleep(5)

    fileName = driver.execute_script("""
        const file_name = document.querySelector('downloads-manager')
                .shadowRoot.getElementById('frb0')
                .shadowRoot.getElementById('file-link').textContent;
        return file_name;
    """)

    data = parse_relatorio_0186(fileName)

    time.sleep(1)

    path = os.path.join('.', fileName)
    if os.path.exists(path):
        os.remove(path)

    for d in data:
        data = (d['data'], d['comanda'], d['item'], d['tipo'], d['categoria'],
                d['profissional'], d['assistente_1'], d['assistente_2'],
                d['comissao_percentual'], d['cliente'], d['email'], d['telefone'],
                d['celular'], d['valor'], d['desconto'], d['quantidade'],
                d['custo'], d['comissao'], d['liquido'], d['ua'])
        
        res = verify_relatorio_0186(data)

        if res != True:
            insert_relatorio_0186(data, idEmpresa)

    time.sleep(4)

def relatorio_0063(driver, idEmpresa):
    global EMPRESA_ID

    driver.get('https://admin.avec.beauty/admin/relatorios/0063')
    time.sleep(10)

    driver.execute_script("document.getElementsByClassName('buttons-excel')[0].click()")

    time.sleep(2)

    driver.get('chrome://downloads')
    time.sleep(4)

    fileName = driver.execute_script("""
        const file_name = document.querySelector('downloads-manager')
                .shadowRoot.getElementById('frb0')
                .shadowRoot.getElementById('file-link').textContent;
        return file_name;
    """)

    try:
        data = parse_relatorio_0063(fileName)
    except Exception as e:
        add_log('Error parsing report 0063: ' + str(e))

    time.sleep(1)

    path = os.path.join('.', fileName)
    if os.path.exists(path):
        os.remove(path)

    try:
        for d in data:
            data = (d['cliente'], d['pacote'], d['servico'], d['total'], d['utilizados'], d['validade'], d['compra'])
            res = verify_relatorio_0063(data)

            if res != True:
                insert_relatorio_0063(data, idEmpresa)
    except Exception as e:
        add_log('Error inserting report 0063: ' + str(e))

    time.sleep(4)

def relatorio_0123(driver, idEmpresa):
    driver.get('https://admin.avec.beauty/admin/relatorios/0123')
    time.sleep(3)

    driver.execute_script("document.getElementsByClassName('inputVarRel')[0].value = '{}'".format(INIT_DATE_0123))
    driver.execute_script("document.getElementsByClassName('inputVarRel')[1].value = '{}'".format(FINAL_DATE_0123))

    time.sleep(15)

    driver.execute_script("document.getElementsByClassName('btn btn-info hidden-print')[0].click()")

    time.sleep(7)

    driver.execute_script("document.getElementsByClassName('buttons-excel')[0].click()")

    time.sleep(2)

    driver.get('chrome://downloads')
    time.sleep(4)

    fileName = driver.execute_script("""
        const file_name = document.querySelector('downloads-manager')
                .shadowRoot.getElementById('frb0')
                .shadowRoot.getElementById('file-link').textContent;
        return file_name;
    """)

    try:
        data = parse_relatorio_0123(fileName)
    except Exception as e:
        add_log('Error parsing report 0123: ' + str(e))

    time.sleep(1)

    path = os.path.join('.', fileName)
    if os.path.exists(path):
        os.remove(path)

    try:
        for d in data:
            _data = (d['profissional'], d['tipo_contratacao'], d['cargo'], d['banco'], d['agencia'], d['conta'], d['faturado'],d['rateio_servicos'], d['rateio_produtos'],d['rateio_outros'], d['caixinha'],d['descontos'], d['a_pagar'], d['valor_casa'])
            res = verify_relatorio_0123(_data)

            if res != True:
                insert_relatorio_0123(_data, idEmpresa)
    except Exception as e:
        add_log('Error inserting report 0123: ' + str(e))
    
    time.sleep(4)

def relatorio_0053(driver, idEmpresa):
    driver.get('https://admin.avec.beauty/admin/relatorios/0053')
    time.sleep(3)

    driver.execute_script("document.getElementsByClassName('inputVarRel')[0].value = '{}'".format(INIT_DATE_0053))
    driver.execute_script("document.getElementsByClassName('inputVarRel')[1].value = '{}'".format(FINAL_DATE_0053))

    time.sleep(15)

    driver.execute_script("document.getElementsByClassName('btn btn-info hidden-print')[0].click()")

    time.sleep(7)

    driver.execute_script("document.getElementsByClassName('buttons-excel')[0].click()")

    time.sleep(2)

    driver.get('chrome://downloads')
    time.sleep(4)

    fileName = driver.execute_script("""
        const file_name = document.querySelector('downloads-manager')
                .shadowRoot.getElementById('frb0')
                .shadowRoot.getElementById('file-link').textContent;
        return file_name;
    """)

    try:
        data = parse_relatorio_0053(fileName)
    except Exception as e:
        add_log('Error parsing report 0053: ' + str(e))

    time.sleep(1)

    path = os.path.join('.', fileName)
    if os.path.exists(path):
        os.remove(path)

    try:
        for d in data:
            data = (d['data_reserva'], d['hora'], d['cliente'], d['servico'], d['valor'])
            res = verify_relatorio_0053(data)

            if res != True:
                insert_relatorio_0053(data, idEmpresa)
    except Exception as e:
        add_log('Error inserting report 0053: ' + str(e))

    time.sleep(4)

def relatorio_0004(driver, idEmpresa):
    driver.get('https://admin.avec.beauty/admin/relatorios/0004')
    time.sleep(10)

    driver.execute_script("document.getElementsByClassName('buttons-excel')[0].click()")

    time.sleep(2)

    driver.get('chrome://downloads')
    time.sleep(4)

    fileName = driver.execute_script("""
        const file_name = document.querySelector('downloads-manager')
                .shadowRoot.getElementById('frb0')
                .shadowRoot.getElementById('file-link').textContent;
        return file_name;
    """)

    try:
        data = parse_relatorio_0004(fileName)
    except Exception as e:
        add_log('Error parsing report 0004: ' + str(e))

    time.sleep(1)

    path = os.path.join('.', fileName)
    if os.path.exists(path):
        os.remove(path)

    try:
        for d in data:
            data = (
                d['cliente'], 
                d['codigo'], 
                d['aniversario'], 
                d['telefone'], 
                d['celular'], 
                d['email'], 
                d['sexo'],
                d['como_conheceu'], 
                d['cpf'], 
                d['cep'], 
                d['endereco'], 
                d['numero'], 
                d['estado'], 
                d['cidade'], 
                d['complemento'], 
                d['bairro'], 
                d['profissao'], 
                d['cadastrado'], 
                d['obs'], 
                d['rg'], 
            )

            res = verify_relatorio_0004(data)

            if res != True:
                insert_relatorio_0004(data, idEmpresa)
    except Exception as e:
        add_log('Error inserting report 0004: ' + str(e))

    time.sleep(4)

def do_scraping(EXECUTING, email, password, link, idEmpresa, startDate, finalDate):
    global EMPRESA_EMAIL
    global EMPRESA_LINK
    global EMPRESA_PASSWORD
    global EMPRESA_ID
    global INIT_DATE_0053
    global INIT_DATE_0123
    global INIT_DATE_0186
    global FINAL_DATE_0053
    global FINAL_DATE_0123
    global FINAL_DATE_0186

    EMPRESA_EMAIL = email
    EMPRESA_PASSWORD = password
    EMPRESA_LINK = link
    EMPRESA_ID = idEmpresa
    INIT_DATE_0186 = startDate
    FINAL_DATE_0186 = finalDate
    INIT_DATE_0123 = startDate
    FINAL_DATE_0123 = finalDate
    INIT_DATE_0053 = startDate
    FINAL_DATE_0053 = finalDate

    driver = Webdriver(os.path.dirname(__file__)).getDriver()

    login(driver=driver)

    relatorio_0063(driver=driver, idEmpresa=idEmpresa)
    relatorio_0123(driver=driver, idEmpresa=idEmpresa)
    relatorio_0004(driver=driver, idEmpresa=idEmpresa)
    relatorio_0053(driver=driver, idEmpresa=idEmpresa)
    relatorio_0186(driver=driver, idEmpresa=idEmpresa)

    driver.quit()

    EXECUTING.append(False)